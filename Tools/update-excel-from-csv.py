#!/usr/bin/env python3
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

# Load CSV
csv_file = Path("languages.csv")
excel_file = Path("languages.xlsx")

if not csv_file.exists():
    print(f"❌ Error: {csv_file} not found")
    exit(1)

# Load existing workbook or create new one
if excel_file.exists():
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    # Clear existing data except header
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Translations"

# Write header
ws['A1'] = "Key"
ws['B1'] = "English"
ws['C1'] = "Vietnamese"

# Write data from CSV
row_num = 2
with open(csv_file, 'r', encoding='utf-8') as f:
    reader = csv.reader(f, delimiter=';')
    next(reader)  # Skip header
    for row in reader:
        if len(row) >= 3:
            ws.cell(row=row_num, column=1).value = row[0]
            ws.cell(row=row_num, column=2).value = row[1]
            ws.cell(row=row_num, column=3).value = row[2]
            row_num += 1

# Adjust column widths
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 50

# Save
wb.save(excel_file)
print(f"✅ Updated {excel_file} from {csv_file}")
